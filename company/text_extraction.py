import re

from docx  import Document
from bs4 import BeautifulSoup
import pypandoc
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import fitz
from xlrd import open_workbook


def read_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def read_docx(file_path):
    doc = Document(file_path)
    text = ' '.join([paragraph.text for paragraph in doc.paragraphs])
    return text


def read_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        contents = file.read()

    soup = BeautifulSoup(contents, 'html.parser')

    # Remove script and style elements
    for script in soup(["script", "style"]):
        script.decompose()

    # Get the text
    text = soup.get_text()

    # Break into lines and remove leading and trailing spaces on each line
    lines = (line.strip() for line in text.splitlines())
    # Break multi-headlines into a line each
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    # Drop blank lines
    text = '\n'.join(chunk for chunk in chunks if chunk)

    return text

def read_xls(file_path):
    try: 
        workbook = load_workbook(filename=file_path, read_only=True)
        words=[]
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and isinstance(cell, str):
                        words_in_cell = re.findall(r'\b\w+\b', cell)
                        words.extend(words_in_cell)
        return ' '.join(words)
    except(InvalidFileException):
        words=[] 
        workbook=open_workbook(filename=file_path)
        for sheet in workbook.sheets(): 
            for ncol in range(sheet.ncols): 
                cells=sheet.col(ncol)
                for cell in cells: 
                    if cell.ctype ==1: 
                        words.extend(cell.value)
        return ' '.join(words)


def read_rtf(file_path): 
    output = pypandoc.convert_file(file_path, 'plain')
    return output




def extracting_text(file_name:str,ext:str): 
    look_up_table_ft={
        'pdf':read_pdf,
        'docx':read_docx,
        'html':read_html,
        'htm':read_html,
        'xls':read_xls,
        'rtf':read_rtf,
    }
    
    result_flag=look_up_table_ft.get(ext,None)
    if result_flag is None: 
        
        print("warning in extracting_text unexpected file type ",ext)
        extracted_text='/n'
    else:
        extracted_text=result_flag(file_name)   
    return extracted_text


